# FUNCTIONS  -----------------------------------------------------------------------------------------------------------------------

def parse_parms(parms_file):
    print('looking for a .parms file at: '+parms_file)
    try:
        with open(parms_file,'r') as inp:
            try:
                parms = eval(inp.read())
            except:
                print("Something wrong with parameters file.")
                raise
        return parms
    except:
        print("Specified parameters file not found.")
        raise
        
def printS(message):
    print(message, end='', flush=True)
    return

# ---------------------------------------------------------------------------------------------------------------------------------
print('Loading libraries...')

import sys
import os
import json
import csv
import glob
from pathlib import Path
import numpy as np
from datetime import datetime
import openpyxl
import requests
import matplotlib.pyplot as plt
from geopandas import read_file, GeoDataFrame
from shapely.geometry import Point, Polygon






# load parameters:
if len(sys.argv) > 1:
    parms = parse_parms(sys.argv[1])
else:
    raise Exception('Not enough input arguments!')




"""
    The function takes as input the paths of the database and Area of interest,
    Return as output a csv file with the target of interest in the AoC.
    
    CRS:
    Every designated target is registered in the database with his own coordinates
    reference system, identified by the EPSG code.
    Its coordinates are transformed into the 'targetCRS' reference system, and 
    compared with the coordinates of the Area of Interest.
    Therefore the AoC coordinates must be provided in the 'targetCRS' reference system.
    (TODO -> add code so that any CRS is accepted for the AoC)
    
"""
# --- CONSTANTS ----------------------

# starting row
cellRowstart = 5

# ID and type columns
cellColID = 4
cellColType = 2

# coordinates columns
cellColEast = 11
cellColNorth = 15
cellColElev = 19

# code columns
cellColEastEPSG = 13
cellColNorthEPSG = 17
cellColElevEPSG = 20

# dates columns
cellColstartDate = 29
cellColendDate = 30

# corrections columns
cellColdxE = 23
cellColdxW = 26

# geometry columns
cellColshape = 38
cellColsize = 38
cellColMRA = 39 # Max Reflecion Azimuth
cellColMRE = 40 # Max Reflecion Elevation

# Target Coordinates Reference System (CRS)
targetCRS = 4258  # ETRS89
CRSname = 'ETRS89'
frameName = 'ETRF2000'
frameEpoch = '2021'
print('****************************************\nfor any inquiry please contact: bazzocchip@gmail.com')


# csv file Head
csvHead = ['ID','TYPE','INSTALLDATE','STARTDATE','ENDDATE',
           'LATITUDE','LONGITUDE','EL.HEIGHT','ORIENTATION',
           'BAND','CRSHAPE','LEGLENGTH','AZIDIP','ZENDIP']


# --------- PARSED PARAMETERS --------
aoiPath = Path(parms["aoiDir"])
dbPath = parms["targetDB"]
outDir = parms["outDir"]
convFlag = parms["convFlag"]
projName = parms["project"]
mapFlag = parms["mapFlag"]
exclTarget = parms["exclTarget"]

print(f'Excluding targets n.: {str(exclTarget)}')

if not(os.path.exists(outDir) and os.path.isdir(outDir)):
    try:
        os.mkdir(outDir)
        print(f"Directory '{outDir}' created successfully.")
    except OSError as error:
        print(f"Error creating directory '{outDir}': {error}")
# ------------------------------------

# Load the Area of Interest
if not os.path.exists(aoiPath):
    raise Exception('The provided shapefile path doesn\'t exists')
shpPath = [p for p in aoiPath.rglob('*interest.shp')]
shpData = read_file(shpPath[0])  
geometry = shpData.geometry
AoC = geometry[0]
gdfList = [{'Type': 'Polygon', 'Name': 'AoC', 'Geometry': AoC}]

# Load the Database
workbook = openpyxl.load_workbook(dbPath)
sheet = workbook['Database']

# Initialize transformed coordinates csv file
csvCoordPath = outDir + 'EPSG_'+str(targetCRS)+'_coordinates.csv'
if convFlag == 1:
    csvCoord = open(csvCoordPath, 'w', newline='')
    csvwriter = csv.writer(csvCoord)
    # Head
    row = ['ID','LAT','EPSG code lat',
            'LON','EPSG code lon',
            'ELEV','EPSG code elev']
    csvwriter.writerow(row)


# Initialize variables for the while loop
selectedTargets = []
content = 'dummy'
trCount = 0

cellRow = cellRowstart


print('\n****************************************')

while (content is not None):
    # for the current row
    
    # check if to exclude
    if (cellRow-cellRowstart+1) in exclTarget:
        print('excluding target from list...')
        # prepare next iteration
        cellRow = cellRow+1
        content = sheet.cell(row=cellRow, column=cellColNorth).value
        continue # The current target is to be excluded a priori
            
    # Extract target ID and type
    targetID = str(sheet.cell(row=cellRow, column=cellColID).value)
    targetType = str(sheet.cell(row=cellRow, column=cellColType).value)
    
    # check if to exclude
    if targetType != 'IGRS':
        print('not an IGRS - excluding target from list...')
        # prepare next iteration
        cellRow = cellRow+1
        content = sheet.cell(row=cellRow, column=cellColNorth).value
        continue # The current target is to be excluded a priori
    
    # Extract coordinates
    Easting = sheet.cell(row=cellRow, column=cellColEast).value
    Northing = sheet.cell(row=cellRow, column=cellColNorth).value
    Elevation = sheet.cell(row=cellRow, column=cellColElev).value
    
    # Extract EPSG code
    epsgEasting = sheet.cell(row=cellRow, column=cellColEastEPSG).value
    epsgNorthing = sheet.cell(row=cellRow, column=cellColNorthEPSG).value
    epsgElevation = sheet.cell(row=cellRow, column=cellColElevEPSG).value
    
    if ( (int(epsgEasting) == targetCRS) and (int(epsgNorthing) == targetCRS ) ):
        #skip the transformation
        lat = float(Northing)
        lon = float(Easting)
        elev = float(Elevation)
    else:
        # Transform CRS via API
        api_url = "http://epsg.io/trans"
        trCount = trCount+1
        params = {
            'x': float(Easting),
            'y': float(Northing),
            'z': float(Elevation),
            's_srs': int(epsgEasting),  # Source CRS (WGS 84)
            't_srs': targetCRS   # Target CRS (ETRS89 / Poland CS92)
        }
        
        response = requests.get(api_url, params=params)

        if response.status_code == 200:
            # Parse the JSON response
            coordinates = response.json()
    
            # Now you can work with the data as needed
            lat = float(coordinates["y"])
            lon = float(coordinates["x"])
            elev = float(coordinates["z"])
        else:
            print("EPSG API request failed with status code:", response.status_code)

    # Save point object
    targetP = Point(lon,lat,elev)
    
    # Save in a csv file all the target only in the targetCRS. This works only if required
    if convFlag == 1:
            row = [targetID,lat,targetCRS,lon,targetCRS,elev,targetCRS]
            csvwriter.writerow(row)

        
    # Check if the target is inside the AoI
    if AoC.contains(targetP): 
        
        # orientation (performed first for convenience)
        ascFlag = ('W' in targetType) or ('IGRS' in targetType) or ('B' in targetType) or ('FF' in targetType)
        dscFlag = ('E' in targetType) or ('IGRS' in targetType) or ('B' in targetType) or ('FF' in targetType)
        band = 'C'
        
        # Display data and switch between type
        RCS0dB = None
        if 'IGRS' in targetType:
            print('\rrow n. '+ str(cellRow-cellRowstart+1)+', IGRS found: '+targetID) 
        elif 'CR' in targetType:
            print('\rrow n. '+ str(cellRow-cellRowstart+1)+', CR found: '+targetID)
            targetType = 'CREF'
        elif 'TR' in targetType:
            print('\rrow n. '+ str(cellRow-cellRowstart+1)+', TR found: '+targetID)
            targetType = 'CAT'
            RCS0dB = 44
        else:
            print('\rrow n. '+ str(cellRow-cellRowstart+1)+', target found: '+targetID)
       
        
        # extract dates
        startDate = str(sheet.cell(row=cellRow, column=cellColstartDate).value)
        installDate = startDate
        endDate = str(sheet.cell(row=cellRow, column=cellColendDate).value)
        
        try:
            installDate = datetime.strptime(installDate, '%Y%m%d').strftime('%Y%m%dT%H%MZ')
        except ValueError:
            installDate = installDate
        
        try:
            startDate = datetime.strptime(startDate, '%Y%m%d').strftime('%Y%m%dT%H%MZ')
        except ValueError:
            if 'TBD' in startDate:
                print('not installed yet, skipping...')
                # prepare next iteration
                cellRow = cellRow+1
                content = sheet.cell(row=cellRow, column=cellColNorth).value
                continue # The current target is not saved in the list because it is not installed yet
            else:
                startDate = startDate
        
        try:
            endDate = datetime.strptime(endDate, '%Y%m%d').strftime('%Y%m%dT%H%MZ')
        except ValueError:
            if ('TODAY' in endDate) or (endDate == 'None'):
                endDate = '99999999T9999Z' 
            else:
                endDate = endDate
        
        
        # Apply Corrections
        dAsc = np.array([0,0,0])
        dDsc = np.array([0,0,0])
        for i in range(0,3):
            try:
                dAsc[i] = float(sheet.cell(row=cellRow, column=cellColdxE+i).value)
            except Exception:
                dAsc[i] = 0
        
        for i in range(0,3):
            try:
                dDsc[i] = float(sheet.cell(row=cellRow, column=cellColdxW+i).value)
            except Exception:
                dDsc[i] = 0
         
        tanB = 0.99664719 * np.tan(np.deg2rad(lat))
        c1 = 6378137*np.pi/180*np.cos(np.arctan(tanB))
        c4 = 111132.954 - 559.882*np.cos(2*np.deg2rad(lat)) + 1.175*np.cos(4*np.deg2rad(lat))
        dConv = np.array([c1,c4,1])
        
        plhAsc = np.array([lon,lat,elev]) + dAsc / dConv
        plhASc = np.array([lon,lat,elev]) + np.array([-3.524*1e-6,-4.208*1e-7,-2.093])
        plhDsc = np.array([lon,lat,elev]) + dDsc / dConv
        plhDSc = np.array([lon,lat,elev]) + np.array([+3.524*1e-6,-4.208*1e-7,-2.093])
        
        # extract geometry
        shape = str(sheet.cell(row=cellRow, column=cellColshape).value)
        leglength = sheet.cell(row=cellRow, column=cellColsize).value
        MRA = sheet.cell(row=cellRow, column=cellColMRA).value
        MRE = sheet.cell(row=cellRow, column=cellColMRE).value
        
                       
        # save into a dictionary
        targetDict = {
            "id" : targetID,
            "type" : targetType,
            "installDate" : installDate,
            "startDate" : startDate,
            "endDate" : endDate,
            "ascending" : {
                "orientation" : ascFlag,
                "band" : band,
                "coordinates":{
                    "longitude": plhAsc[0],
                    "latitude" : plhAsc[1],
                    "elevation": plhAsc[2],
                    "CRS" : CRSname,
                    "FRAME" : frameName,
                    "EPOCH": frameEpoch,
                    "EPSG": targetCRS
                },
                "RCS0" : None
            },
            "descending" : {
                "orientation" : dscFlag,
                "band" : band,
                "coordinates":{
                    "longitude": plhDsc[0],
                    "latitude" : plhDsc[1],
                    "elevation": plhDsc[2],
                    "CRS" : CRSname,
                    "FRAME" : frameName,
                    "EPOCH": frameEpoch,
                    "EPSG": targetCRS
                },
                "RCS0" : None
            },
            "geometry":{
                "shape": targetType,#shape,
                "flipped": True,
                "leglength": leglength,
                "azimuthDip": MRA,
                "zenithDip": MRE
            },
            "RCS0": RCS0dB
        }

        selectedTargets.append(targetDict)

        # Append to geodataframe
        gdfList.append({'Type': 'Point', 'Name': targetID+': '+targetType, 'Geometry': targetP})        
        
    else:
        printS('\rrow n. '+ str(cellRow-cellRowstart+1))
    
    
    # prepare next iteration
    cellRow = cellRow+1
    content = sheet.cell(row=cellRow, column=cellColNorth).value        
    

print('\r'+str(len(selectedTargets)) + ' target in the Area of Interest...')

csvCoord.close()
print(f'\nDONE: EPSG:{targetCRS} coordinates exported at {csvCoordPath}, '+str(trCount)+' coords transformed')

# export the json file
stPath = outDir + 'reflectors.json'

with open(stPath, 'w') as json_file:
    json.dump({"stations" : selectedTargets}, json_file)
    print(f'DONE: json list exported to {stPath}')
    
    
    
    
# export the CSV file
csvpath = outDir + 'reflectors.csv'
with open(csvpath, 'w', newline='') as csvfile:
    csvwriter = csv.writer(csvfile)
    csvwriter.writerow(csvHead)
    for target in selectedTargets:
        row = [target["id"],target["type"],
               target["installDate"],target["startDate"],target["endDate"],
               target["ascending"]["coordinates"]["latitude"],
               target["ascending"]["coordinates"]["longitude"],
               target["ascending"]["coordinates"]["elevation"],
               target["ascending"]["orientation"],
               target["ascending"]["band"],
               target["geometry"]["shape"],
               target["geometry"]["leglength"],
               target["geometry"]["azimuthDip"],
               target["geometry"]["zenithDip"]]
        csvwriter.writerow(row)
    print(f'DONE: CSV input file for GECORIS exported at {csvpath}')


if mapFlag == 1:
    
    gdf = GeoDataFrame(gdfList, geometry='Geometry',crs='EPSG:'+str(targetCRS))
    
    gdf.to_file(outDir+'SelectedTargetsMap.geojson', driver="GeoJSON")
    
    print(f'DONE: geojson file exported at '+outDir+'SelectedTargetsMap.geojson\n      visit geojson.io to display it')
    print('TODO: the shape of the target is currently the targetTypoe --> fix it')